"""Componentes de la interfaz de usuario Streamlit.

Contiene todas las funciones de renderizado de la aplicacion web:
configuracion de pagina, barra lateral de parametros, pestanas de
prediccion, grupos, descripcion del modelo, integracion LLM y editor
de datos.  El punto de entrada principal es ``main()`` llamado desde
``app.py``.
"""

from __future__ import annotations

import html
import importlib
import json
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from dotenv import load_dotenv

from wcup2026.config import (
    APP_CAPTION,
    APP_TITLE,
    CHART_COLORS,
    DATA_PATH,
    FIFA_GROUPS_URL,
    FIFA_SCHEDULE_URL,
    GROUPS,
    OPENAI_RESPONSES_URL,
)
from wcup2026.data import (
    apply_ratings_update,
    dataframe_from_csv_text,
    dataframe_to_csv_text,
    load_team_data,
    validate_team_data,
)
from wcup2026.llm import (
    api_key_available,
    build_analysis_payload,
    call_llm_analysis,
    call_llm_group_results_update,
    call_llm_knockout_results_update,
    call_llm_news_search,
    call_llm_ratings_update,
    default_model,
)
from wcup2026.parameters import SimParams
from wcup2026.persistence import (
    POST_GROUP_STATE_PATH,
    load_bracket,
    load_bracket_probable,
    load_llm_analysis,
    load_montecarlo_results,
    load_post_group_state,
    save_llm_analysis,
    save_montecarlo_results,
    save_post_group_state,
)
import wcup2026.report as report_module
from wcup2026.simulator import (
    build_knockout_state_from_group_results,
    describe_matchup,
    simulate_bracket_from_group_results,
    simulate_bracket_most_probable,
    simulate_bracket_most_probable_from_group_results,
    simulate_bracket_sample,
    simulate_knockout_projection_from_group_results,
    simulate_many,
)


def configure_page() -> None:
    """Configurar la pagina Streamlit y cargar variables de entorno.

    Llama a ``st.set_page_config`` con titulo e icono de la app, carga
    el archivo ``.env`` via dotenv e inyecta los estilos CSS personalizados.
    Debe invocarse como primera instruccion Streamlit del script.

    Returns
    -------
    None
        Configura Streamlit y no devuelve ningun valor.
    """
    st.set_page_config(page_title=APP_TITLE, page_icon="WC26", layout="wide")
    load_dotenv()
    inject_style()


def inject_style() -> None:
    """Inyectar CSS personalizado en la pagina via ``st.markdown``.

    Ajusta padding del contenedor principal, estilo de las metricas,
    tipografia de encabezados y clases utilitarias ``.source-line`` y
    ``.small-note``.

    Returns
    -------
    None
        Escribe estilos CSS en la pagina actual de Streamlit.
    """
    st.markdown(
        """
        <style>
        .block-container {padding-top: 1.25rem; padding-bottom: 2rem;}
        [data-testid="stMetric"] {
            border: 1px solid #3a3a3a;
            border-radius: 8px;
            padding: 12px 14px;
        }
        h1, h2, h3 {letter-spacing: 0;}
        .source-line {
            color: #5d6759;
            font-size: 0.88rem;
        }
        .small-note {
            color: #5d6759;
            font-size: 0.92rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def apply_plotly_theme(
    fig: go.Figure,
    *,
    height: int | None = None,
    showlegend: bool | None = None,
) -> go.Figure:
    """Aplicar estilo visual consistente a figuras Plotly de la app."""
    layout_updates = {
        "paper_bgcolor": "rgba(0,0,0,0)",
        "plot_bgcolor": "rgba(0,0,0,0)",
        "font": {"color": "#1f2937", "family": "Arial, sans-serif"},
        "margin": {"l": 10, "r": 18, "t": 42, "b": 32},
        "hoverlabel": {"bgcolor": "#111827", "font": {"color": "#f9fafb", "size": 12}},
    }
    if height is not None:
        layout_updates["height"] = height
    if showlegend is not None:
        layout_updates["showlegend"] = showlegend
    fig.update_layout(**layout_updates)
    fig.update_xaxes(
        showgrid=True,
        gridcolor="rgba(148, 163, 184, 0.22)",
        zeroline=False,
        linecolor="rgba(148, 163, 184, 0.35)",
    )
    fig.update_yaxes(
        showgrid=False,
        zeroline=False,
        linecolor="rgba(148, 163, 184, 0.35)",
    )
    return fig


def render_copy_button(text: str, key: str) -> None:
    """Renderizar un boton HTML para copiar texto al portapapeles del navegador.

    Inyecta un componente HTML con JavaScript que llama a
    ``navigator.clipboard.writeText`` al hacer clic.  El ``key`` debe ser
    unico en la pagina para evitar colisiones de IDs.

    Parameters
    ----------
    text : str
        Contenido que se copiara al portapapeles al pulsar el boton.
    key : str
        Sufijo unico que se usa para construir los IDs del boton y el
        indicador de estado en el DOM.
    """
    button_id = f"copy-llm-{key}"
    status_id = f"copy-llm-status-{key}"
    payload = json.dumps(text)
    button_label = html.escape("Copiar analisis")
    st.iframe(
            f"""
            <div style="display:flex; justify-content:flex-end; margin:0 0 0.5rem 0;">
                <button
                    id="{button_id}"
                    type="button"
                    style="border:1px solid #d0d7de; border-radius:0.5rem; background:#f6f8fa; color:#24292f; padding:0.4rem 0.8rem; font-size:0.9rem; cursor:pointer;"
                >
                    {button_label}
                </button>
                <span id="{status_id}" style="margin-left:0.5rem; font-size:0.85rem; color:#57606a;"></span>
            </div>
            <script>
            const copyButton = document.getElementById({json.dumps(button_id)});
            const status = document.getElementById({json.dumps(status_id)});
            const text = {payload};

            copyButton?.addEventListener('click', async () => {{
                try {{
                    await navigator.clipboard.writeText(text);
                    if (status) {{
                        status.textContent = 'Copiado';
                        setTimeout(() => {{ status.textContent = ''; }}, 2000);
                    }}
                }} catch (error) {{
                    if (status) {{
                        status.textContent = 'No se pudo copiar';
                    }}
                }}
            }});
            </script>
            """,
            height=42,
    )


@st.cache_data(show_spinner=False)
def load_default_data_cached() -> pd.DataFrame:
    """Cargar los datos de equipos predeterminados con cache de Streamlit.

    Returns
    -------
    pd.DataFrame
        DataFrame leido desde ``DATA_PATH``; el resultado queda cacheado
        para evitar lecturas de disco repetidas.
    """
    return load_team_data()


@st.cache_data(show_spinner=False)
def run_bracket_sample_cached(csv_text: str, params: SimParams) -> pd.DataFrame:
    """Generar el cuadro de eliminacion de una simulacion representativa con cache.

    Envuelve ``simulate_bracket_sample`` con cache de Streamlit.  El cache
    se invalida cuando cambia ``csv_text`` o ``params``.

    Parameters
    ----------
    csv_text : str
        Datos de equipos serializados como CSV (hace hashable el DataFrame).
    params : SimParams
        Hiperparametros de la simulacion.

    Returns
    -------
    pd.DataFrame
        Partidos de eliminatoria (salida de ``simulate_bracket_sample``).
    """
    df = dataframe_from_csv_text(csv_text)
    validate_team_data(df)
    return simulate_bracket_sample(df, params)


@st.cache_data(show_spinner=False)
def run_bracket_probable_cached(csv_text: str, params: SimParams) -> pd.DataFrame:
    """Generar el cuadro mas probable con cache.

    Envuelve ``simulate_bracket_most_probable`` con cache de Streamlit.  El
    cache se invalida cuando cambia ``csv_text`` o ``params``.

    Parameters
    ----------
    csv_text : str
        Datos de equipos serializados como CSV (hace hashable el DataFrame).
    params : SimParams
        Hiperparametros de la simulacion.

    Returns
    -------
    pd.DataFrame
        Cuadro mas probable calculado sobre 1000 simulaciones (salida de
        ``simulate_bracket_most_probable``).
    """
    df = dataframe_from_csv_text(csv_text)
    validate_team_data(df)
    return simulate_bracket_most_probable(df, params, n=1000)


@st.cache_data(show_spinner=False)
def run_post_group_bracket_cached(
    csv_text: str,
    group_results_csv_text: str,
    knockout_results_csv_text: str,
    knockout_input_csv_text: str,
    params: SimParams,
) -> pd.DataFrame:
    """Generar una llave representativa desde resultados finales de grupos."""
    df = dataframe_from_csv_text(csv_text)
    group_results = dataframe_from_csv_text(group_results_csv_text)
    knockout_results = dataframe_from_csv_text(knockout_results_csv_text)
    knockout_input = dataframe_from_csv_text(knockout_input_csv_text)
    validate_team_data(df)
    return simulate_bracket_from_group_results(
        df,
        group_results,
        params,
        knockout_results=knockout_results,
        r32_fixtures=knockout_input,
    )


@st.cache_data(show_spinner=False)
def run_post_group_bracket_probable_cached(
    csv_text: str,
    group_results_csv_text: str,
    knockout_results_csv_text: str,
    knockout_input_csv_text: str,
    params: SimParams,
) -> pd.DataFrame:
    """Generar la llave mas probable desde resultados finales de grupos."""
    df = dataframe_from_csv_text(csv_text)
    group_results = dataframe_from_csv_text(group_results_csv_text)
    knockout_results = dataframe_from_csv_text(knockout_results_csv_text)
    knockout_input = dataframe_from_csv_text(knockout_input_csv_text)
    validate_team_data(df)
    return simulate_bracket_most_probable_from_group_results(
        df,
        group_results,
        params,
        n=1000,
        knockout_results=knockout_results,
        r32_fixtures=knockout_input,
    )


@st.cache_data(show_spinner=False)
def run_post_group_projection_cached(
    csv_text: str,
    group_results_csv_text: str,
    knockout_results_csv_text: str,
    knockout_input_csv_text: str,
    params: SimParams,
) -> pd.DataFrame:
    """Estimar probabilidades de eliminatorias desde grupos ya definidos."""
    df = dataframe_from_csv_text(csv_text)
    group_results = dataframe_from_csv_text(group_results_csv_text)
    knockout_results = dataframe_from_csv_text(knockout_results_csv_text)
    knockout_input = dataframe_from_csv_text(knockout_input_csv_text)
    validate_team_data(df)
    return simulate_knockout_projection_from_group_results(
        df,
        group_results,
        params,
        knockout_results=knockout_results,
        r32_fixtures=knockout_input,
    )


@st.cache_data(show_spinner=False)
def build_post_group_knockout_state_cached(
    csv_text: str,
    group_results_csv_text: str,
    knockout_results_csv_text: str,
    knockout_input_csv_text: str,
    params: SimParams,
) -> pd.DataFrame:
    """Construir estado editable de eliminatorias desde grupos + resultados KO parciales."""
    df = dataframe_from_csv_text(csv_text)
    group_results = dataframe_from_csv_text(group_results_csv_text)
    knockout_results = dataframe_from_csv_text(knockout_results_csv_text)
    knockout_input = dataframe_from_csv_text(knockout_input_csv_text)
    validate_team_data(df)
    return build_knockout_state_from_group_results(
        df,
        group_results,
        params,
        knockout_results=knockout_results,
        r32_fixtures=knockout_input,
    )


@st.cache_data(show_spinner="Simulando torneos...")
def run_simulation_cached(csv_text: str, params: SimParams) -> pd.DataFrame:
    """Ejecutar la simulacion con cache de Streamlit.

    El cache se invalida cuando cambia ``csv_text`` o ``params``, por lo
    que cualquier edicion en los ratings o los controles laterales
    desencadena una nueva simulacion.

    Parameters
    ----------
    csv_text : str
        Datos de equipos serializados como CSV (hace hashable el DataFrame).
    params : SimParams
        Hiperparametros de la simulacion.

    Returns
    -------
    pd.DataFrame
        Resultados de la simulacion (salida de ``simulate_many``).
    """
    df = dataframe_from_csv_text(csv_text)
    validate_team_data(df)
    return simulate_many(df, params)


def build_results_export_csv(results: pd.DataFrame) -> str:
    """Construir el CSV de probabilidades finales en formato portable.

    Convierte las columnas internas en porcentaje a probabilidades 0-1 y
    conserva una fila por cada seleccion simulada.

    Parameters
    ----------
    results : pd.DataFrame
        DataFrame de resultados de la simulacion (salida de
        ``simulate_many``) con columnas ``team``, ``champion_pct``,
        ``final_pct`` y ``semifinal_pct``.

    Returns
    -------
    str
        Texto CSV con columnas ``team``, ``prob_champion``, ``prob_final``
        y ``prob_semifinal`` (probabilidades 0-1 con cuatro decimales).
    """
    export = results[
        ["team", "champion_pct", "final_pct", "semifinal_pct"]
    ].rename(
        columns={
            "champion_pct": "prob_champion",
            "final_pct": "prob_final",
            "semifinal_pct": "prob_semifinal",
        }
    ).copy()
    for column in ["prob_champion", "prob_final", "prob_semifinal"]:
        export[column] = export[column] / 100
    return export.to_csv(index=False, float_format="%.4f")


def load_persisted_outputs_once() -> None:
    """Cargar resultados guardados en disco una sola vez por sesion de Streamlit.

    Usa la clave ``_persisted_outputs_loaded`` del ``st.session_state`` como
    centinela para evitar lecturas repetidas en re-runs.  Inicializa
    ``simulation_results``, ``simulation_df`` y ``llm_answer`` si existen
    datos guardados y aun no hay valores en la sesion.

    Returns
    -------
    None
        Actualiza ``st.session_state`` con datos persistidos si existen.
    """
    if st.session_state.get("_persisted_outputs_loaded"):
        return

    try:
        persisted = load_montecarlo_results()
        if persisted is not None:
            results, teams = persisted
            st.session_state.setdefault("simulation_results", results)
            st.session_state.setdefault("simulation_df", teams)
    except Exception as exc:  # pragma: no cover - UI guardrail
        st.warning(f"No se pudo cargar la simulacion guardada: {exc}")

    try:
        bracket = load_bracket()
        if bracket is not None:
            st.session_state.setdefault("bracket", bracket)
    except Exception as exc:  # pragma: no cover - UI guardrail
        st.warning(f"No se pudo cargar el cuadro guardado: {exc}")

    try:
        bracket_probable = load_bracket_probable()
        if bracket_probable is not None:
            st.session_state.setdefault("bracket_probable", bracket_probable)
    except Exception as exc:  # pragma: no cover - UI guardrail
        st.warning(f"No se pudo cargar el cuadro mas probable guardado: {exc}")

    try:
        llm_answer = load_llm_analysis()
        if llm_answer:
            st.session_state.setdefault("llm_answer", llm_answer)
    except Exception as exc:  # pragma: no cover - UI guardrail
        st.warning(f"No se pudo cargar el analisis LLM guardado: {exc}")

    st.session_state.setdefault(
        "_post_group_state_file_available",
        POST_GROUP_STATE_PATH.exists(),
    )

    st.session_state["_persisted_outputs_loaded"] = True


def render_sidebar() -> tuple[SimParams, str]:
    """Renderizar la barra lateral con todos los controles del motor.

    Expone sliders y campos para numero de simulaciones, semilla, goles
    base, ventaja de anfitrion, ruido en eliminatorias y pesos de los
    componentes del rating.  Tambien muestra la configuracion del LLM.

    Returns
    -------
    tuple[SimParams, str]
        ``(params, model)`` donde ``params`` es el ``SimParams`` construido
        con los valores de los controles y ``model`` es el nombre del
        modelo OpenAI configurado.
    """
    st.sidebar.header("Motor")
    simulations = st.sidebar.slider("Simulaciones", 500, 50000, 5000, step=500, help="Numero de torneos simulados via Monte Carlo. Mas simulaciones = mayor precision, pero mas lento.")
    seed = st.sidebar.number_input("Semilla", min_value=1, value=2026, step=1, help="Semilla aleatoria para reproducibilidad. El mismo valor siempre genera los mismos resultados.")
    base_goals = st.sidebar.slider("Goles base por equipo", 0.8, 2.2, 1.35, step=0.05, help="Promedio de goles esperados por equipo en un partido completamente neutral. Ajusta el ritmo ofensivo global del torneo.")
    home_advantage = st.sidebar.slider("Ventaja anfitrion", 0.0, 0.25, 0.10, step=0.01, help="Multiplicador adicional de goles para los equipos anfitriones (USA, Canada, Mexico). 0.10 = +10% de goles esperados.")
    knockout_noise = st.sidebar.slider("Ruido en eliminatorias", 8.0, 30.0, 18.0, step=0.5, help="Factor de aleatoriedad en partidos eliminatorios. Valores altos favorecen sorpresas; valores bajos privilegian al mejor equipo.")

    st.sidebar.header("Pesos")
    elo_weight = st.sidebar.slider("Elo / ranking", 0.0, 1.0, 0.45, step=0.05, help="Peso del Elo/ranking FIFA en el rating compuesto de cada seleccion. Refleja historial y nivel competitivo acumulado.")
    squad_weight = st.sidebar.slider("Plantilla", 0.0, 1.0, 0.25, step=0.05, help="Peso de la calidad del plantel (valor de mercado, profundidad). Equipos con mejor banco aguantan mejor el torneo.")
    form_weight = st.sidebar.slider("Forma reciente", 0.0, 1.0, 0.15, step=0.05, help="Peso de los resultados recientes (ultimos 6-12 meses). Captura momentum y dinamica actual del equipo.")
    balance_weight = st.sidebar.slider("Ataque + defensa", 0.0, 1.0, 0.15, step=0.05, help="Peso del balance tactico (ataque vs defensa). Equipos con ratings altos en ambos extremos son mas solidos.")

    st.sidebar.header("LLM")
    model = st.sidebar.text_input("Modelo OpenAI", value=default_model(), help="Nombre del modelo de OpenAI a usar para el analisis cualitativo y busqueda de noticias. Ej: gpt-4o, gpt-5.5.")
    key_status = "detectada" if api_key_available() else "no detectada"
    st.sidebar.caption(f"OPENAI_API_KEY: {key_status}")

    params = SimParams(
        simulations=int(simulations),
        seed=int(seed),
        base_goals=float(base_goals),
        home_advantage=float(home_advantage),
        knockout_noise=float(knockout_noise),
        elo_weight=float(elo_weight),
        squad_weight=float(squad_weight),
        form_weight=float(form_weight),
        balance_weight=float(balance_weight),
    )
    return params, model


def render_probability_view(results: pd.DataFrame) -> None:
    """Renderizar la pestana de prediccion con metricas, grafico y tabla.

    Muestra cuatro metricas del favorito, un grafico de barras horizontales
    con los 12 equipos mas probables campeones coloreados por confederacion
    y una tabla completa con todas las probabilidades por etapa.

    Parameters
    ----------
    results : pd.DataFrame
        DataFrame de resultados ordenado por probabilidad de campeon
        (salida de ``simulate_many``).
    """
    top = results.head(12).copy()
    favorite = results.iloc[0]
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Favorito", favorite["team"], f"{favorite['champion_pct']:.1f}% campeon", help="Equipo con mayor probabilidad de ganar el torneo segun la simulacion Monte Carlo.")
    col2.metric("Final", favorite["team"], f"{favorite['final_pct']:.1f}%", help="Probabilidad de que el favorito llegue a la final del torneo.")
    col3.metric("Semifinal", favorite["team"], f"{favorite['semifinal_pct']:.1f}%", help="Probabilidad de que el favorito alcance al menos las semifinales.")
    col4.metric("Rating modelo", f"{favorite['overall']:.1f}", "0-100 relativo", help="Rating compuesto del favorito en escala 0-100. Combina Elo, plantilla, forma y balance tactico segun los pesos configurados.")

    stage_cols = [
        "round_of_32_pct",
        "round_of_16_pct",
        "quarterfinal_pct",
        "semifinal_pct",
        "final_pct",
        "champion_pct",
    ]
    stage_labels = ["R32", "Octavos", "Cuartos", "Semis", "Final", "Campeon"]

    tab_rank, tab_path, tab_heatmap = st.tabs(["Favoritos", "Ruta por ronda", "Mapa de avance"])

    with tab_rank:
        sorted_top = top.sort_values("champion_pct")
        fig = go.Figure()
        fig.add_trace(
            go.Bar(
                x=sorted_top["champion_pct"],
                y=sorted_top["team"],
                orientation="h",
                marker={
                    "color": sorted_top["champion_pct"],
                    "colorscale": [[0, "#dbeafe"], [0.45, "#60a5fa"], [1, "#14532d"]],
                    "line": {"color": "rgba(31, 41, 55, 0.22)", "width": 0.7},
                },
                text=sorted_top["champion_pct"].map(lambda value: f"{value:.1f}%"),
                textposition="outside",
                customdata=sorted_top[["confederation", "final_pct", "semifinal_pct", "overall"]],
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Confederacion: %{customdata[0]}<br>"
                    "Campeon: %{x:.1f}%<br>"
                    "Final: %{customdata[1]:.1f}%<br>"
                    "Semifinal: %{customdata[2]:.1f}%<br>"
                    "Rating: %{customdata[3]:.1f}<extra></extra>"
                ),
            )
        )
        fig.update_layout(title="Top 12 por probabilidad de campeon")
        fig.update_xaxes(title="Probabilidad de campeon (%)", ticksuffix="%")
        fig.update_yaxes(title="")
        fig.update_traces(cliponaxis=False)
        st.plotly_chart(apply_plotly_theme(fig, height=520, showlegend=False), width="stretch")

    with tab_path:
        route_top = results.head(8).copy()
        fig = go.Figure()
        palette = CHART_COLORS + px.colors.qualitative.Set2
        for idx, row in route_top.reset_index(drop=True).iterrows():
            values = [row[col] for col in stage_cols]
            fig.add_trace(
                go.Scatter(
                    x=stage_labels,
                    y=values,
                    mode="lines+markers",
                    name=row["team"],
                    line={"width": 3, "color": palette[idx % len(palette)]},
                    marker={"size": 8},
                    hovertemplate=(
                        f"<b>{row['team']}</b><br>"
                        "%{x}: %{y:.1f}%<extra></extra>"
                    ),
                )
            )
        fig.update_layout(title="Probabilidad acumulada de avanzar por ronda", legend={"orientation": "h", "y": -0.22})
        fig.update_yaxes(title="Probabilidad (%)", range=[0, 105], ticksuffix="%")
        fig.update_xaxes(title="")
        st.plotly_chart(apply_plotly_theme(fig, height=500), width="stretch")

    with tab_heatmap:
        heat = results.head(16).copy()
        z_values = heat[stage_cols].round(1).to_numpy()
        text_values = [[f"{value:.1f}%" for value in row] for row in z_values]
        fig = go.Figure(
            go.Heatmap(
                z=z_values,
                x=stage_labels,
                y=heat["team"],
                text=text_values,
                texttemplate="%{text}",
                colorscale=[[0, "#f8fafc"], [0.35, "#93c5fd"], [0.7, "#22c55e"], [1, "#14532d"]],
                colorbar={"title": "%"},
                hovertemplate="<b>%{y}</b><br>%{x}: %{z:.1f}%<extra></extra>",
            )
        )
        fig.update_layout(title="Mapa de avance del top 16")
        fig.update_xaxes(side="top")
        fig.update_yaxes(autorange="reversed")
        st.plotly_chart(apply_plotly_theme(fig, height=560, showlegend=False), width="stretch")

    st.download_button(
        "Exportar resultados CSV",
        data=build_results_export_csv(results),
        file_name="resultados_wcup2026.csv",
        mime="text/csv",
        help="Descarga todas las selecciones con columnas team, prob_champion, prob_final y prob_semifinal en escala 0-1.",
    )

    st.dataframe(
        results[
            [
                "team",
                "group",
                "confederation",
                "overall",
                "group_winner_pct",
                "round_of_32_pct",
                "round_of_16_pct",
                "quarterfinal_pct",
                "semifinal_pct",
                "final_pct",
                "champion_pct",
            ]
        ].round(2),
        width="stretch",
        hide_index=True,
    )


def render_group_view(df: pd.DataFrame, results: pd.DataFrame) -> None:
    """Renderizar la pestana de analisis por grupo.

    Permite seleccionar un grupo, muestra la tabla de ratings y
    probabilidades de sus equipos, y ofrece un comparador de duelo
    directo entre dos selecciones del grupo con metricas de xG y
    probabilidades de victoria/empate.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame original de equipos con ratings.
    results : pd.DataFrame
        DataFrame de resultados de la simulacion.
    """
    group = st.selectbox("Grupo", sorted(df["group"].unique()), help="Selecciona el grupo para ver la tabla de ratings y el comparador de duelos directos.")
    merged = df.merge(
        results[["team", "overall", "round_of_32_pct", "champion_pct"]],
        on="team",
        how="left",
    )
    view = merged.loc[merged["group"] == group].sort_values("overall", ascending=False)
    st.dataframe(
        view[
            [
                "team",
                "confederation",
                "elo",
                "attack",
                "defense",
                "squad",
                "form",
                "overall",
                "round_of_32_pct",
                "champion_pct",
            ]
        ].round(2),
        width="stretch",
        hide_index=True,
    )

    chart_col, radar_col = st.columns([1.2, 1])
    with chart_col:
        fig = go.Figure()
        marker_sizes = (view["champion_pct"].fillna(0) + 4).clip(lower=8, upper=42)
        fig.add_trace(
            go.Scatter(
                x=view["attack"],
                y=view["defense"],
                mode="markers+text",
                text=view["team"],
                textposition="top center",
                marker={
                    "size": marker_sizes,
                    "color": view["round_of_32_pct"].fillna(0),
                    "colorscale": [[0, "#e0f2fe"], [0.5, "#38bdf8"], [1, "#166534"]],
                    "showscale": True,
                    "colorbar": {"title": "R32 %"},
                    "line": {"color": "#1f2937", "width": 0.7},
                },
                customdata=view[["overall", "champion_pct", "confederation"]],
                hovertemplate=(
                    "<b>%{text}</b><br>"
                    "Ataque: %{x:.1f}<br>"
                    "Defensa: %{y:.1f}<br>"
                    "Rating: %{customdata[0]:.1f}<br>"
                    "Campeon: %{customdata[1]:.1f}%<br>"
                    "%{customdata[2]}<extra></extra>"
                ),
            )
        )
        fig.update_layout(title=f"Perfil competitivo del Grupo {group}")
        fig.update_xaxes(title="Ataque", range=[max(0, view["attack"].min() - 8), min(100, view["attack"].max() + 8)])
        fig.update_yaxes(title="Defensa", range=[max(0, view["defense"].min() - 8), min(100, view["defense"].max() + 8)])
        st.plotly_chart(apply_plotly_theme(fig, height=420, showlegend=False), width="stretch")

    with radar_col:
        radar_metrics = ["attack", "defense", "squad", "form", "overall"]
        radar_labels = ["Ataque", "Defensa", "Plantilla", "Forma", "Rating"]
        fig = go.Figure()
        for idx, row in view.reset_index(drop=True).iterrows():
            values = [row[col] for col in radar_metrics]
            fig.add_trace(
                go.Scatterpolar(
                    r=values + [values[0]],
                    theta=radar_labels + [radar_labels[0]],
                    fill="toself",
                    name=row["team"],
                    opacity=0.68,
                    line={"color": (CHART_COLORS + px.colors.qualitative.Set2)[idx % (len(CHART_COLORS) + len(px.colors.qualitative.Set2))]},
                    hovertemplate=f"<b>{row['team']}</b><br>%{{theta}}: %{{r:.1f}}<extra></extra>",
                )
            )
        fig.update_layout(
            title="Radar de ratings",
            polar={
                "radialaxis": {
                    "visible": True,
                    "range": [0, 100],
                    "gridcolor": "rgba(148, 163, 184, 0.28)",
                },
                "bgcolor": "rgba(0,0,0,0)",
            },
            legend={"orientation": "h", "y": -0.18},
        )
        st.plotly_chart(apply_plotly_theme(fig, height=420), width="stretch")

    teams = view["team"].tolist()
    col1, col2 = st.columns(2)
    team_a = col1.selectbox("Equipo A", teams, index=0, help="Primera seleccion del duelo directo simulado.")
    team_b = col2.selectbox("Equipo B", teams, index=min(1, len(teams) - 1), help="Segunda seleccion del duelo directo simulado.")
    if team_a != team_b:
        params = st.session_state["params"]
        matchup = describe_matchup(team_a, team_b, df, params, samples=6000)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric(f"Gana {team_a}", f"{matchup['team_a_win_pct']:.1f}%", help=f"Probabilidad de victoria de {team_a} en 6000 simulaciones del partido.")
        c2.metric("Empate", f"{matchup['draw_pct']:.1f}%", help="Probabilidad de empate al final del tiempo reglamentario.")
        c3.metric(f"Gana {team_b}", f"{matchup['team_b_win_pct']:.1f}%", help=f"Probabilidad de victoria de {team_b} en 6000 simulaciones del partido.")
        c4.metric("xG", f"{matchup['team_a_xg']:.2f} - {matchup['team_b_xg']:.2f}", help="Goles esperados promedio (xG) para cada equipo segun sus ratings de ataque y defensa.")
        matchup_fig = go.Figure(
            go.Bar(
                x=[matchup["team_a_win_pct"], matchup["draw_pct"], matchup["team_b_win_pct"]],
                y=[f"Gana {team_a}", "Empate", f"Gana {team_b}"],
                orientation="h",
                marker={"color": ["#166534", "#94a3b8", "#1d4ed8"]},
                text=[
                    f"{matchup['team_a_win_pct']:.1f}%",
                    f"{matchup['draw_pct']:.1f}%",
                    f"{matchup['team_b_win_pct']:.1f}%",
                ],
                textposition="outside",
                hovertemplate="%{y}: %{x:.1f}%<extra></extra>",
            )
        )
        matchup_fig.update_layout(title=f"Distribucion del duelo: {team_a} vs {team_b}")
        matchup_fig.update_xaxes(title="Probabilidad (%)", range=[0, 100], ticksuffix="%")
        matchup_fig.update_yaxes(title="")
        matchup_fig.update_traces(cliponaxis=False)
        st.plotly_chart(apply_plotly_theme(matchup_fig, height=300, showlegend=False), width="stretch")


def render_model_view() -> None:
    """Renderizar la pestana de descripcion del modelo.

    Muestra texto explicativo sobre el funcionamiento del simulador,
    criterios de desempate, logica de eliminatorias y enlaces a las
    fuentes de datos oficiales de FIFA y OpenAI.

    Returns
    -------
    None
        Renderiza contenido informativo en la pagina Streamlit.
    """
    st.subheader("Como piensa el modelo")
    st.markdown(
        """
        - Cada seleccion tiene rating de ataque, defensa, plantilla, forma y Elo aproximado.
        - Cada partido se simula con goles Poisson a partir de ataque vs defensa y rating global.
        - La fase de grupos usa puntos, diferencia de goles, goles a favor y rating como desempate final.
        - Avanzan dos primeros de cada grupo y los ocho mejores terceros.
        - La ronda de 32 sigue el calendario FIFA; la asignacion de terceros usa las ventanas de grupos publicadas por FIFA.
        - Las eliminatorias se deciden por marcador simulado; si hay empate, se usa una probabilidad de prorroga/penales basada en rating.
        """
    )
    st.markdown(
        f"""
        <div class="source-line">
        Fuentes base:
        <a href="{FIFA_GROUPS_URL}" target="_blank">grupos FIFA</a>,
        <a href="{FIFA_SCHEDULE_URL}" target="_blank">calendario FIFA</a>,
        <a href="{OPENAI_RESPONSES_URL}" target="_blank">OpenAI Responses API</a>.
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.info(
        "Los ratings del CSV son semillas editables, no una verdad oficial. La gracia del MVP es que puedas calibrarlos con datos mejores: Elo actualizado, lesiones, valor de mercado, minutos jugados y resultados recientes."
    )


def render_llm_view(
    results: pd.DataFrame,
    df: pd.DataFrame,
    model: str,
    bracket_probable: pd.DataFrame | None = None,
) -> None:
    """Renderizar la pestana de integracion LLM.

    Muestra ideas de uso del LLM, un area de texto para ingresar
    informacion cualitativa y un boton para generar el analisis.  Si la
    clave de API no esta disponible muestra una advertencia en su lugar.

    Parameters
    ----------
    results : pd.DataFrame
        DataFrame de resultados de la simulacion.
    df : pd.DataFrame
        DataFrame original de equipos con ratings.
    model : str
        Nombre del modelo OpenAI a usar en la llamada.
    """
    st.subheader("LLM en la ecuacion")

    if not api_key_available():
        st.warning("No detecte OPENAI_API_KEY en el entorno. Revisa que el archivo .env este en esta carpeta.")
        return

    if st.button("Buscar noticias", help="Consulta al LLM por lesiones y contexto relevante de los equipos"):
        try:
            with st.spinner("Buscando noticias relevantes..."):
                st.session_state["llm_notes"] = call_llm_news_search(model, df)
        except Exception as exc:  # pragma: no cover - UI guardrail
            st.error(f"No se pudo obtener noticias: {exc}")

    notes = st.text_area(
        "Escenario o informacion cualitativa",
        placeholder="Ejemplo: Francia llega sin su lateral titular; Mexico juega con presion local; Uruguay trae buena racha defensiva.",
        height=180,
        key="llm_notes",
        help="Agrega contexto que el modelo cuantitativo no captura: lesiones, suspensiones, clima, rivalidades, rotaciones o escenarios hipoteticos. Puedes editarlo libremente despues de usar 'Buscar noticias'.",
    )

    if st.button("Generar analisis LLM", type="primary", help="Envia los resultados de la simulacion y el escenario cualitativo al LLM para obtener un analisis integrado con recomendaciones accionables."):
        if results is None:
            st.warning("Primero pulsa **Simular torneo** para tener resultados que analizar.")
        else:
            payload = build_analysis_payload(
                results, df, notes,
                bracket_probable=(
                    bracket_probable
                    if bracket_probable is not None
                    else st.session_state.get("bracket_probable")
                ),
            )
            try:
                with st.spinner("Consultando al LLM..."):
                    st.session_state["llm_answer"] = call_llm_analysis(model, payload).strip()
                    save_llm_analysis(st.session_state["llm_answer"])
            except Exception as exc:  # pragma: no cover - UI guardrail
                st.error(f"No se pudo consultar el LLM: {exc}")

    if "llm_answer" in st.session_state:
        answer = st.session_state["llm_answer"]
        if answer:
            render_copy_button(answer, key="analysis")
            st.markdown(answer)
        else:
            st.warning("El LLM respondio sin texto visible. Revisa el modelo configurado o intenta de nuevo.")


def render_report_view(
    results: pd.DataFrame | None,
    df: pd.DataFrame,
    bracket_probable: pd.DataFrame | None = None,
) -> None:
    """Renderizar la pestana de generacion del reporte LaTeX/PDF.

    Muestra un boton para generar el archivo TEX y compilar el PDF.  Si
    aun no hay simulacion realizada, muestra un mensaje informativo.
    Tras la compilacion exitosa ofrece botones de descarga para el PDF
    y el TEX.

    Parameters
    ----------
    results : pd.DataFrame or None
        DataFrame de resultados de la simulacion.  Si es ``None``, se
        muestra un aviso pidiendo simular primero.
    df : pd.DataFrame
        DataFrame original de equipos con ratings.
    """
    st.subheader("Reporte")
    st.markdown(
        '<div class="small-note">Genera reporte/reporte_wcup2026.tex desde el template con placeholders y compila el PDF con pdflatex.</div>',
        unsafe_allow_html=True,
    )

    if results is None:
        st.info("Pulsa **Simular torneo** antes de generar el reporte.")
        return

    if st.button("Generar reporte", type="primary", help="Crea el archivo LaTeX final y compila reporte_wcup2026.pdf con pdflatex."):
        try:
            with st.spinner("Generando reporte LaTeX y compilando PDF..."):
                fresh_report_module = importlib.reload(report_module)
                tex_path, pdf_path = fresh_report_module.generate_report(
                    results=results,
                    teams=df,
                    llm_text=st.session_state.get("llm_answer"),
                    params=st.session_state.get("params"),
                    bracket_probable=(
                        bracket_probable
                        if bracket_probable is not None
                        else st.session_state.get("bracket_probable")
                    ),
                    compile_pdf=True,
                )
            st.success(f"Reporte generado: {tex_path.name}")
            if pdf_path is not None:
                st.caption(f"PDF compilado: {pdf_path}")
                st.download_button(
                    "Descargar PDF",
                    data=pdf_path.read_bytes(),
                    file_name=pdf_path.name,
                    mime="application/pdf",
                )
            st.download_button(
                "Descargar TEX",
                data=tex_path.read_text(encoding="utf-8"),
                file_name=tex_path.name,
                mime="application/x-tex",
            )
        except Exception as exc:
            st.error(f"No se pudo generar el reporte: {exc}")


def render_data_editor(default_df: pd.DataFrame, model: str) -> pd.DataFrame:
    """Renderizar el editor de ratings con actualizacion via IA.

    Muestra un boton para obtener ratings actualizados usando busqueda web
    a traves del LLM, y permite editar directamente los valores de ataque,
    defensa, plantilla y forma en una tabla interactiva.

    Parameters
    ----------
    default_df : pd.DataFrame
        DataFrame predeterminado que se muestra si no se ha realizado
        ninguna actualizacion via IA.
    model : str
        Nombre del modelo OpenAI a usar para la actualizacion de ratings.

    Returns
    -------
    pd.DataFrame
        DataFrame con los ratings (posiblemente editados) que se usa para
        la simulacion.
    """
    if "working_df" not in st.session_state:
        st.session_state["working_df"] = default_df.copy()

    if api_key_available():
        if st.button(
            "Actualizar ratings con IA",
            help="Usa busqueda web para obtener ratings actualizados (Elo, forma, plantilla) de todas las selecciones y reemplaza los valores actuales.",
        ):
            try:
                with st.spinner("Buscando datos actualizados con IA..."):
                    updates = call_llm_ratings_update(model, st.session_state["working_df"])
                    st.session_state["working_df"] = apply_ratings_update(
                        st.session_state["working_df"], updates
                    )
                st.session_state["working_df"].to_csv(DATA_PATH, index=False)
                load_default_data_cached.clear()
                st.success(f"Ratings actualizados para {len(updates)} selecciones y guardados en {DATA_PATH.name}.")
                st.rerun()
            except Exception as exc:
                st.error(f"No se pudo actualizar los ratings: {exc}")
    else:
        st.caption("OPENAI_API_KEY no detectada. No es posible actualizar ratings con IA.")

    with st.expander("Editar ratings del modelo", expanded=False):
        st.markdown(
            '<div class="small-note">Edita valores de 0 a 100 para ataque, defensa, plantilla y forma. Luego vuelve a correr la simulacion.</div>',
            unsafe_allow_html=True,
        )
        return st.data_editor(
            st.session_state["working_df"],
            width="stretch",
            num_rows="fixed",
            column_config={
                "is_host": st.column_config.CheckboxColumn("is_host", help="Marca si el equipo es uno de los tres anfitriones del torneo (USA, Canada, Mexico)."),
                "attack": st.column_config.NumberColumn("attack", min_value=0, max_value=100, help="Potencial ofensivo del equipo (0-100). Afecta los goles esperados generados."),
                "defense": st.column_config.NumberColumn("defense", min_value=0, max_value=100, help="Solidez defensiva del equipo (0-100). Reduce los goles esperados del rival."),
                "squad": st.column_config.NumberColumn("squad", min_value=0, max_value=100, help="Calidad y profundidad del plantel (0-100). Equipos con mejor banco aguantan mejor torneos largos."),
                "form": st.column_config.NumberColumn("form", min_value=0, max_value=100, help="Forma reciente del equipo (0-100). Refleja los resultados de los ultimos 6-12 meses."),
            },
        )


def _build_bracket_figure(bracket: pd.DataFrame, from_round: str = "round_of_16"):
    """Construir la figura Plotly del cuadro de eliminacion como arbol vertical.

    Distribuye cada partido en columnas y filas fijas, dibuja conexiones
    entre rondas, resalta al ganador y agrega marcadores invisibles para
    tooltips Plotly.  Si el bracket esta vacio, los partidos se renderizan
    como TBD.

    Parameters
    ----------
    bracket : pd.DataFrame
        DataFrame con columnas ``round``, ``match_id``, ``team_a``,
        ``team_b``, ``winner`` y opcionalmente ``winner_pct``.
    from_round : str, optional
        ``"round_of_32"`` para mostrar todo el cuadro, ``"round_of_16"``
        para mostrar desde octavos o ``"quarterfinal"`` para mostrar solo
        desde cuartos.
        Por defecto ``"round_of_16"``.

    Returns
    -------
    plotly.graph_objects.Figure
        Figura Plotly lista para renderizar en Streamlit.
    """
    import plotly.graph_objects as go

    round_labels = {
        "round_of_32": "Ronda de 32",
        "round_of_16": "Octavos",
        "quarterfinal": "Cuartos",
        "semifinal": "Semifinal",
        "final": "Final",
    }

    if from_round == "round_of_32":
        r32_order = [74, 77, 73, 75, 83, 84, 81, 82, 76, 78, 79, 80, 86, 88, 85, 87]
        positions = {match_id: (0, 30.0 - idx * 2.0) for idx, match_id in enumerate(r32_order)}
        positions.update({
            89: (1, 29.0), 90: (1, 25.0),
            93: (1, 21.0), 94: (1, 17.0),
            91: (1, 13.0), 92: (1, 9.0),
            95: (1, 5.0), 96: (1, 1.0),
            97: (2, 27.0), 98: (2, 19.0),
            99: (2, 11.0), 100: (2, 3.0),
            101: (3, 23.0), 102: (3, 7.0),
            104: (4, 15.0),
        })
        col_x = {0: 0.0, 1: 3.6, 2: 7.2, 3: 10.8, 4: 14.4}
        col_labels = {
            0: "Ronda de 32",
            1: "Octavos",
            2: "Cuartos",
            3: "Semifinales",
            4: "Final",
        }
        box_hh = 0.42
        y_range = [-1.7, 32.2]
        x_range = [-1.7, 16.2]
        fig_height = 920
        font_size = 8
    elif from_round == "quarterfinal":
        positions = {
            97: (0, 9.0), 98: (0, 6.0),
            99: (0, 3.0), 100: (0, 0.0),
            101: (1, 7.5), 102: (1, 1.5),
            104: (2, 4.5),
        }
        col_x = {0: 0.0, 1: 3.5, 2: 7.0}
        col_labels = {0: "Cuartos de Final", 1: "Semifinales", 2: "Final"}
        box_hh = 0.65
        y_range = [-1.5, 11.5]
        x_range = [-1.5, 9.0]
        fig_height = 500
        font_size = 11
    else:
        positions = {
            89: (0, 14.0), 90: (0, 12.0),
            93: (0, 10.0), 94: (0, 8.0),
            91: (0, 6.0),  92: (0, 4.0),
            95: (0, 2.0),  96: (0, 0.0),
            97: (1, 13.0), 98: (1, 9.0),
            99: (1, 5.0),  100: (1, 1.0),
            101: (2, 11.0), 102: (2, 3.0),
            104: (3, 7.0),
        }
        col_x = {0: 0.0, 1: 3.5, 2: 7.0, 3: 10.5}
        col_labels = {0: "Octavos", 1: "Cuartos", 2: "Semifinales", 3: "Final"}
        box_hh = 0.45
        y_range = [-1.5, 16.5]
        x_range = [-1.5, 12.5]
        fig_height = 700
        font_size = 9

    parents = {
        89: (74, 77), 90: (73, 75),
        91: (76, 78), 92: (79, 80),
        93: (83, 84), 94: (81, 82),
        95: (86, 88), 96: (85, 87),
        97: (89, 90), 98: (93, 94),
        99: (91, 92), 100: (95, 96),
        101: (97, 98), 102: (99, 100),
        104: (101, 102),
    }
    box_hw = 1.45
    match_dict = bracket.set_index("match_id").to_dict("index") if not bracket.empty else {}

    shapes = []
    annotations = []
    hover_x: list[float] = []
    hover_y: list[float] = []
    hover_text: list[str] = []

    def clean_value(value) -> str:
        """Normalizar valores vacios de pandas para etiquetas."""
        if value is None:
            return ""
        text = str(value)
        return "" if text in ("None", "nan", "NaN") else text

    def compact_label(text: str, max_chars: int) -> str:
        """Acortar etiquetas largas sin perder legibilidad dentro de cajas."""
        text = clean_value(text)
        if len(text) <= max_chars:
            return text
        return text[: max_chars - 1].rstrip() + "…"

    def match_round(match_id: int) -> str:
        """Inferir nombre de ronda a partir del ID de partido."""
        if 73 <= match_id <= 88:
            return "round_of_32"
        if 89 <= match_id <= 96:
            return "round_of_16"
        if 97 <= match_id <= 100:
            return "quarterfinal"
        if 101 <= match_id <= 102:
            return "semifinal"
        return "final"

    for child_id, (p1_id, p2_id) in parents.items():
        if child_id not in positions or p1_id not in positions or p2_id not in positions:
            continue
        child_col, child_y = positions[child_id]
        p1_col, p1_y = positions[p1_id]
        _p2_col, p2_y = positions[p2_id]
        px_right = col_x[p1_col] + box_hw
        cx_left = col_x[child_col] - box_hw
        conn_x = (px_right + cx_left) / 2
        child_row = match_dict.get(child_id, {})
        child_teams = {clean_value(child_row.get("team_a")), clean_value(child_row.get("team_b"))}
        for py in (p1_y, p2_y):
            parent_id = p1_id if py == p1_y else p2_id
            parent_winner = clean_value(match_dict.get(parent_id, {}).get("winner"))
            line_color = "#4ecca3" if parent_winner and parent_winner in child_teams else "#4b5563"
            line_width = 2.4 if line_color == "#4ecca3" else 1.2
            shapes.append({"type": "line", "x0": px_right, "y0": py, "x1": conn_x, "y1": py,
                           "line": {"color": line_color, "width": line_width}})
        shapes.append({"type": "line", "x0": conn_x, "y0": p1_y, "x1": conn_x, "y1": p2_y,
                       "line": {"color": "#4b5563", "width": 1.2}})
        shapes.append({"type": "line", "x0": conn_x, "y0": child_y, "x1": cx_left, "y1": child_y,
                       "line": {"color": "#4b5563", "width": 1.2}})

    for match_id, (col, y) in positions.items():
        cx = col_x[col]
        if match_id in match_dict:
            row = match_dict[match_id]
            team_a = clean_value(row["team_a"])
            team_b = clean_value(row["team_b"])
            winner = clean_value(row["winner"])
            raw_pct = row.get("winner_pct")
            winner_pct = float(raw_pct) if raw_pct is not None and str(raw_pct) not in ("None", "nan", "") else None
        else:
            team_a, team_b, winner, winner_pct = "TBD", "TBD", "", None

        for slot_y0, slot_y1, team in [
            (y, y + box_hh * 2, team_a),
            (y - box_hh * 2, y, team_b),
        ]:
            is_winner = team == winner and team not in ("", "TBD")
            fill = "#14532d" if is_winner else "#1f2937"
            border = "#4ade80" if is_winner else "#475569"
            shapes.append({"type": "rect",
                           "x0": cx - box_hw, "y0": slot_y0,
                           "x1": cx + box_hw, "y1": slot_y1,
                           "fillcolor": fill,
                           "line": {"color": border, "width": 1.3 if is_winner else 0.8}})
            color = "#bbf7d0" if is_winner else "#e5e7eb"
            label = compact_label(team, 18 if from_round == "round_of_32" else 22)
            if is_winner and winner_pct is not None:
                label = compact_label(f"{team} ({winner_pct:.0f}%)", 18 if from_round == "round_of_32" else 22)
            slot_mid = (slot_y0 + slot_y1) / 2
            annotations.append({"x": cx, "y": (slot_y0 + slot_y1) / 2,
                                 "text": label, "showarrow": False,
                                 "font": {"color": color, "size": font_size},
                                 "xanchor": "center", "yanchor": "middle"})
            pct_text = f"<br>Frecuencia ganador: {winner_pct:.1f}%" if is_winner and winner_pct is not None else ""
            hover_x.append(cx)
            hover_y.append(slot_mid)
            hover_text.append(
                "<b>"
                + round_labels.get(match_round(match_id), match_round(match_id))
                + f" · Partido {match_id}</b><br>"
                + f"{team_a} vs {team_b}<br>"
                + f"Ganador: {winner or 'TBD'}"
                + pct_text
            )

    label_y = y_range[1] - 0.4
    for col, label in col_labels.items():
        annotations.append({"x": col_x[col], "y": label_y,
                             "text": f"<b>{label}</b>", "showarrow": False,
                             "font": {"color": "#aaaaaa", "size": 11},
                             "xanchor": "center", "yanchor": "top"})

    champion = clean_value(match_dict.get(104, {}).get("winner"))
    if champion and 104 in positions:
        final_x = col_x[positions[104][0]]
        final_y = positions[104][1]
        annotations.append({
            "x": final_x,
            "y": final_y - 1.25,
            "text": f"<b>Campeon: {html.escape(champion)}</b>",
            "showarrow": False,
            "font": {"color": "#facc15", "size": 13},
            "xanchor": "center",
            "yanchor": "middle",
        })

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=hover_x,
            y=hover_y,
            mode="markers",
            marker={"size": 18, "color": "rgba(255,255,255,0.01)"},
            hovertext=hover_text,
            hoverinfo="text",
            showlegend=False,
        )
    )
    fig.update_layout(
        shapes=shapes,
        annotations=annotations,
        paper_bgcolor="#0e1117",
        plot_bgcolor="#0e1117",
        xaxis={"visible": False, "range": x_range},
        yaxis={"visible": False, "range": y_range},
        height=fig_height,
        margin={"l": 10, "r": 10, "t": 28, "b": 10},
        hoverlabel={"bgcolor": "#111827", "font": {"color": "#f9fafb", "size": 12}},
    )
    fig.update_xaxes(fixedrange=True)
    fig.update_yaxes(fixedrange=True)
    return fig


def render_bracket_view(
    bracket: pd.DataFrame | None,
    bracket_probable: pd.DataFrame | None,
    *,
    empty_message: str | None = None,
    probable_label: str = "Mas probable (1000 sim.)",
    sample_label: str = "Una simulacion",
    probable_caption: str | None = None,
    sample_caption: str | None = None,
) -> None:
    """Renderizar el cuadro de eliminacion del torneo como grafico tipo arbol.

    Muestra los partidos desde octavos o cuartos de final hasta la final.
    Permite elegir entre una simulacion representativa (semilla fija) o el
    cuadro mas probable calculado sobre 1000 simulaciones, donde el
    ganador de cada posicion es el equipo que gano con mas frecuencia y
    se muestra su porcentaje de victorias.

    Parameters
    ----------
    bracket : pd.DataFrame or None
        Cuadro de una simulacion representativa.
    bracket_probable : pd.DataFrame or None
        Cuadro mas probable calculado sobre N simulaciones.
    """
    any_data = bracket is not None or bracket_probable is not None
    if not any_data:
        st.info(
            empty_message
            or "Pulsa **Simular torneo** para generar el cuadro de eliminacion. "
            "Los resultados se guardaran en el XLSX para la proxima sesion."
        )
        return

    col_mode, col_round = st.columns([2, 2])
    with col_mode:
        mode_options = []
        if bracket_probable is not None:
            mode_options.append(probable_label)
        if bracket is not None:
            mode_options.append(sample_label)
        mode = st.radio("Modo", mode_options, horizontal=True, key="bracket_mode")
    with col_round:
        start_round = st.radio(
            "Mostrar desde",
            ["Ronda de 32", "Octavos de Final", "Cuartos de Final"],
            horizontal=True,
            key="bracket_start_round",
        )

    from_round = {
        "Ronda de 32": "round_of_32",
        "Octavos de Final": "round_of_16",
        "Cuartos de Final": "quarterfinal",
    }[start_round]
    if mode == probable_label:
        active = bracket_probable
        st.caption(
            probable_caption
            or "Cuadro mas probable: para cada posicion se muestra el equipo que gano con mayor frecuencia en 1000 simulaciones. El porcentaje indica la frecuencia del ganador en esa posicion."
        )
    else:
        active = bracket
        st.caption(
            sample_caption
            or "Simulacion representativa (semilla fija). El ganador de cada partido aparece en verde."
        )

    fig = _build_bracket_figure(active, from_round)
    st.plotly_chart(fig, width='stretch')


def build_group_results_template(df: pd.DataFrame) -> pd.DataFrame:
    """Crear plantilla editable para cargar resultados finales de grupos."""
    clean = df.copy()
    clean["group"] = clean["group"].astype(str).str.upper().str.strip()
    rows = []
    for group in GROUPS:
        group_teams = clean.loc[clean["group"] == group, "team"].tolist()
        for position, team in enumerate(group_teams, start=1):
            rows.append({
                "group": group,
                "position": position,
                "team": team,
                "points": 0,
                "gf": 0,
                "ga": 0,
            })
    return pd.DataFrame(rows)


def normalize_group_results_update(update: pd.DataFrame, df: pd.DataFrame) -> pd.DataFrame:
    """Validar y ordenar una tabla de grupos obtenida por busqueda web."""
    required = ["group", "position", "team", "points", "gf", "ga"]
    missing = set(required).difference(update.columns)
    if missing:
        raise ValueError(f"Faltan columnas: {', '.join(sorted(missing))}")

    valid_pairs = (
        df[["team", "group"]]
        .assign(group=lambda data: data["group"].astype(str).str.upper().str.strip())
        .set_index("team")["group"]
        .to_dict()
    )
    clean = update[required].copy()
    clean["team"] = clean["team"].astype(str).str.strip()
    clean["group"] = clean["group"].astype(str).str.upper().str.strip()
    for column in ["position", "points", "gf", "ga"]:
        clean[column] = pd.to_numeric(clean[column], errors="coerce").fillna(0).astype(int)

    unknown = sorted(set(clean["team"]) - set(valid_pairs))
    if unknown:
        raise ValueError(f"Equipos no reconocidos: {', '.join(unknown)}")
    missing_teams = sorted(set(valid_pairs) - set(clean["team"]))
    if missing_teams:
        raise ValueError(f"Faltan equipos: {', '.join(missing_teams)}")
    duplicated = clean.loc[clean["team"].duplicated(), "team"].tolist()
    if duplicated:
        raise ValueError(f"Equipos duplicados: {', '.join(duplicated)}")

    mismatched = [
        f"{row.team} ({row.group} vs {valid_pairs[row.team]})"
        for row in clean.itertuples()
        if valid_pairs[row.team] != row.group
    ]
    if mismatched:
        raise ValueError("Grupos inconsistentes: " + ", ".join(mismatched))

    invalid_positions = clean.loc[~clean["position"].between(1, 4)]
    if not invalid_positions.empty:
        raise ValueError("Todas las posiciones deben estar entre 1 y 4.")
    duplicate_positions = clean.duplicated(subset=["group", "position"], keep=False)
    if duplicate_positions.any():
        bad = clean.loc[duplicate_positions, ["group", "position"]].drop_duplicates()
        labels = [f"{row.group}-{row.position}" for row in bad.itertuples()]
        raise ValueError(f"Posiciones duplicadas por grupo: {', '.join(labels)}")

    return clean.sort_values(["group", "position"]).reset_index(drop=True)


def normalize_knockout_results_update(update: pd.DataFrame, fixtures: pd.DataFrame) -> pd.DataFrame:
    """Validar ganadores cargados para partidos de eliminatoria."""
    required = {"match_id", "winner"}
    missing = required.difference(update.columns)
    if missing:
        raise ValueError(f"Faltan columnas en resultados KO: {', '.join(sorted(missing))}")

    base_required = {"match_id", "team_a", "team_b"}
    missing_base = base_required.difference(fixtures.columns)
    if missing_base:
        raise ValueError(
            "La tabla de fixtures no incluye columnas requeridas: "
            + ", ".join(sorted(missing_base))
        )

    clean = update[["match_id", "winner"]].copy()
    clean["match_id"] = pd.to_numeric(clean["match_id"], errors="coerce")
    clean = clean.dropna(subset=["match_id"])
    clean["match_id"] = clean["match_id"].astype(int)
    clean["winner"] = clean["winner"].where(clean["winner"].notna(), "")
    clean["winner"] = clean["winner"].astype(str).str.strip()
    clean = clean.loc[~clean["winner"].isin(["", "nan", "None"])] .copy()

    duplicated = clean.loc[clean["match_id"].duplicated(), "match_id"].tolist()
    if duplicated:
        raise ValueError(
            "Partidos duplicados en resultados KO: "
            + ", ".join(map(str, sorted(set(duplicated))))
        )

    fixture_map = (
        fixtures[["match_id", "team_a", "team_b"]]
        .copy()
        .assign(match_id=lambda data: pd.to_numeric(data["match_id"], errors="coerce").fillna(0).astype(int))
        .set_index("match_id")
        .to_dict(orient="index")
    )

    unknown_matches = sorted(set(clean["match_id"]) - set(fixture_map))
    if unknown_matches:
        raise ValueError(
            "Hay partidos KO no reconocidos: " + ", ".join(map(str, unknown_matches))
        )

    errors: list[str] = []
    for row in clean.itertuples():
        fixture = fixture_map[row.match_id]
        team_a = str(fixture["team_a"]).strip()
        team_b = str(fixture["team_b"]).strip()
        if team_a.startswith("Ganador ") or team_b.startswith("Ganador "):
            errors.append(
                f"Partido {row.match_id}: aun no se conocen ambos rivales ({team_a} vs {team_b})."
            )
            continue
        if row.winner not in {team_a, team_b}:
            errors.append(
                f"Partido {row.match_id}: ganador '{row.winner}' no coincide con {team_a} vs {team_b}."
            )
    if errors:
        raise ValueError("; ".join(errors[:6]))

    return clean.sort_values("match_id").reset_index(drop=True)


def save_post_group_state_from_session() -> None:
    """Persistir el estado post-grupos actual del session_state."""
    group_results_input = st.session_state.get("post_group_results_input")
    if not isinstance(group_results_input, pd.DataFrame):
        return

    knockout_input = st.session_state.get("post_group_knockout_input")
    if not isinstance(knockout_input, pd.DataFrame):
        knockout_input = pd.DataFrame(columns=["round", "match_id", "team_a", "team_b", "winner"])

    knockout_results = st.session_state.get("post_group_knockout_results")
    if not isinstance(knockout_results, pd.DataFrame):
        knockout_results = pd.DataFrame(columns=["match_id", "winner"])

    save_post_group_state(
        group_results_input=group_results_input,
        knockout_input=knockout_input,
        knockout_results=knockout_results,
        projection=st.session_state.get("post_group_projection"),
        bracket=st.session_state.get("post_group_bracket"),
        bracket_probable=st.session_state.get("post_group_bracket_probable"),
    )


def clear_post_group_outputs() -> None:
    """Limpiar salidas derivadas cuando cambian los resultados reales."""
    for key in ["post_group_projection", "post_group_bracket", "post_group_bracket_probable"]:
        st.session_state.pop(key, None)


def render_post_group_knockout_view(
    df: pd.DataFrame,
    params: SimParams,
    model: str,
    *,
    show_header: bool = True,
    show_outputs: bool = True,
) -> None:
    """Renderizar modelado de eliminatorias condicionado a grupos reales."""
    if show_header:
        st.subheader("Llaves desde resultados de grupos")
        st.markdown(
            '<div class="small-note">Carga la tabla final de grupos y, opcionalmente, fija resultados reales de eliminatorias por ronda. El modelo vuelve a correr solo para los partidos restantes.</div>',
            unsafe_allow_html=True,
        )
        st.info(
            "Paso 1: valida grupos. Paso 2: registra resultados KO ya jugados (manual o IA). "
            "Paso 3: pulsa Modelar llaves con estos resultados."
        )

    state_file_exists = POST_GROUP_STATE_PATH.exists()
    saved_post_group_state: dict[str, pd.DataFrame] | None = None
    state_file_stamp = "missing"
    if not state_file_exists:
        st.warning(f"No existe {POST_GROUP_STATE_PATH}.")
    else:
        try:
            saved_post_group_state = load_post_group_state()
            state_file_stamp = str(POST_GROUP_STATE_PATH.stat().st_mtime_ns)
            if saved_post_group_state is None:
                st.warning(f"{POST_GROUP_STATE_PATH} no contiene estado post-grupos utilizable.")
            else:
                st.caption("Las tablas usan resultados/post_group_state.xlsx como fuente de verdad.")
        except Exception as exc:
            st.error(f"No se pudo cargar {POST_GROUP_STATE_PATH}: {exc}")

    source_mode = "xlsx" if saved_post_group_state is not None else "manual"
    source_signature = f"post-group-source:{source_mode}:{state_file_stamp}\n" + dataframe_to_csv_text(
        df[["team", "group"]].sort_values(["group", "team"]).reset_index(drop=True)
    )
    if (
        "post_group_results_input" not in st.session_state
        or st.session_state.get("_post_group_results_source") != source_signature
    ):
        if saved_post_group_state is not None and "group_results_input" in saved_post_group_state:
            st.session_state["post_group_results_input"] = saved_post_group_state["group_results_input"].copy()
            st.session_state["post_group_knockout_results"] = saved_post_group_state.get(
                "knockout_results",
                pd.DataFrame(columns=["match_id", "winner"]),
            ).copy()
            st.session_state["post_group_knockout_input"] = saved_post_group_state.get(
                "knockout_input",
                pd.DataFrame(),
            ).copy()
            for state_key, sheet_name in [
                ("post_group_projection", "projection"),
                ("post_group_bracket", "bracket"),
                ("post_group_bracket_probable", "bracket_probable"),
            ]:
                if sheet_name in saved_post_group_state:
                    st.session_state[state_key] = saved_post_group_state[sheet_name].copy()
                else:
                    st.session_state.pop(state_key, None)
        else:
            st.session_state["post_group_results_input"] = build_group_results_template(df)
            st.session_state["post_group_knockout_results"] = pd.DataFrame(columns=["match_id", "winner"])
            st.session_state["post_group_knockout_input"] = pd.DataFrame()
            st.session_state.pop("post_group_projection", None)
            st.session_state.pop("post_group_bracket", None)
            st.session_state.pop("post_group_bracket_probable", None)
        st.session_state["_post_group_results_source"] = source_signature

    with st.expander("Cargar tabla final de fase de grupos", expanded=True):
        st.markdown(
            '<div class="small-note">Reordena equipos por posicion y completa puntos, goles a favor y goles en contra. Si usas el archivo guardado, este editor refleja la fuente de verdad seleccionada arriba.</div>',
            unsafe_allow_html=True,
        )
        if api_key_available():
            if st.button(
                "Actualizar grupos con busqueda web",
                help="Usa OpenAI web_search para buscar standings/resultados recientes de fase de grupos y llenar esta tabla.",
            ):
                try:
                    with st.spinner("Buscando resultados de grupos con OpenAI web_search..."):
                        update = call_llm_group_results_update(model, df)
                        normalized_update = normalize_group_results_update(update, df)
                        st.session_state["post_group_results_input"] = normalized_update

                        current_knockout_results = st.session_state.get(
                            "post_group_knockout_results",
                            pd.DataFrame(columns=["match_id", "winner"]),
                        )
                        if not isinstance(current_knockout_results, pd.DataFrame):
                            current_knockout_results = pd.DataFrame(columns=["match_id", "winner"])

                        current_knockout_input = st.session_state.get(
                            "post_group_knockout_input",
                            pd.DataFrame(columns=["match_id", "team_a", "team_b"]),
                        )
                        if not isinstance(current_knockout_input, pd.DataFrame):
                            current_knockout_input = pd.DataFrame(columns=["match_id", "team_a", "team_b"])

                        csv_text = dataframe_to_csv_text(df)
                        group_csv_text = dataframe_to_csv_text(normalized_update)
                        knockout_csv_text = dataframe_to_csv_text(current_knockout_results)
                        fixture_override_csv_text = dataframe_to_csv_text(current_knockout_input)

                        try:
                            rebuilt_fixtures = build_post_group_knockout_state_cached(
                                csv_text,
                                group_csv_text,
                                knockout_csv_text,
                                fixture_override_csv_text,
                                params,
                            )
                        except ValueError:
                            rebuilt_fixtures = build_post_group_knockout_state_cached(
                                csv_text,
                                group_csv_text,
                                dataframe_to_csv_text(pd.DataFrame(columns=["match_id", "winner"])),
                                dataframe_to_csv_text(pd.DataFrame(columns=["match_id", "team_a", "team_b"])),
                                params,
                            )

                        st.session_state["post_group_knockout_input"] = rebuilt_fixtures
                        st.session_state["post_group_knockout_results"] = normalize_knockout_results_update(
                            rebuilt_fixtures,
                            rebuilt_fixtures,
                        )
                        clear_post_group_outputs()
                        save_post_group_state_from_session()
                    st.success("Tabla de grupos actualizada con busqueda web y guardada en post_group_state.xlsx.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"No se pudieron actualizar los grupos con busqueda web: {exc}")
        else:
            st.caption("OPENAI_API_KEY no detectada. Puedes cargar la tabla manualmente.")

        group_results_input = st.data_editor(
            st.session_state["post_group_results_input"],
            width="stretch",
            num_rows="fixed",
            key="post_group_results_editor",
            disabled=["group", "position"],
            column_config={
                "group": st.column_config.TextColumn("Grupo"),
                "position": st.column_config.NumberColumn("Posicion", min_value=1, max_value=4),
                "team": st.column_config.SelectboxColumn(
                    "Equipo",
                    options=sorted(df["team"].tolist()),
                    required=True,
                ),
                "points": st.column_config.NumberColumn("Pts", min_value=0, max_value=9, step=1),
                "gf": st.column_config.NumberColumn("GF", min_value=0, max_value=30, step=1),
                "ga": st.column_config.NumberColumn("GC", min_value=0, max_value=30, step=1),
            },
        )
        st.session_state["post_group_results_input"] = group_results_input.copy()

    normalized_groups: pd.DataFrame | None = None
    group_error: str | None = None
    try:
        normalized_groups = normalize_group_results_update(group_results_input, df)
    except Exception as exc:
        group_error = str(exc)

    if normalized_groups is None:
        st.warning(
            "Corrige la tabla de grupos para habilitar el modelado de eliminatorias. "
            f"Detalle: {group_error}"
        )
    else:
        prior_ko = st.session_state.get("post_group_knockout_results")
        if prior_ko is None or not isinstance(prior_ko, pd.DataFrame):
            prior_ko = pd.DataFrame(columns=["match_id", "winner"])
        if prior_ko.empty:
            knockout_csv_text = pd.DataFrame(columns=["match_id", "winner"]).to_csv(index=False)
        else:
            knockout_csv_text = dataframe_to_csv_text(prior_ko)

        try:
            if saved_post_group_state is not None and "knockout_input" in saved_post_group_state:
                fixtures_state = saved_post_group_state["knockout_input"].copy()
            else:
                csv_text = dataframe_to_csv_text(df)
                group_csv_text = dataframe_to_csv_text(normalized_groups)
                fixture_override_input = pd.DataFrame(columns=["match_id", "team_a", "team_b"])
                fixture_override_csv_text = dataframe_to_csv_text(fixture_override_input)
                fixtures_state = build_post_group_knockout_state_cached(
                    csv_text,
                    group_csv_text,
                    knockout_csv_text,
                    fixture_override_csv_text,
                    params,
                )
            st.session_state["post_group_knockout_input"] = fixtures_state
        except Exception as exc:
            st.error(f"No se pudo construir el estado de eliminatorias: {exc}")

    fixtures_preview = st.session_state.get("post_group_knockout_input", pd.DataFrame())
    confirmed_count = 0
    if isinstance(fixtures_preview, pd.DataFrame) and not fixtures_preview.empty and "winner" in fixtures_preview.columns:
        confirmed_count = int(fixtures_preview["winner"].astype(str).str.strip().ne("").sum())

    s1, s2, s3 = st.columns(3)
    s1.metric("Estado grupos", "Valido" if normalized_groups is not None else "Invalido")
    s2.metric("Resultados KO cargados", str(confirmed_count))
    s3.metric("Partidos KO totales", "31")

    with st.expander("Resultados reales de eliminatorias (opcional)", expanded=False):
        st.markdown(
            '<div class="small-note">Registra ganadores de partidos ya jugados. El simulador fija esos resultados y vuelve a calcular probabilidades para lo que falta.</div>',
            unsafe_allow_html=True,
        )
        fixtures_input = st.session_state.get("post_group_knockout_input", pd.DataFrame())

        if normalized_groups is None or fixtures_input.empty:
            st.info("Primero deja valida la tabla de grupos para habilitar este editor.")
        else:
            if api_key_available():
                if st.button(
                    "Actualizar eliminatorias con busqueda web",
                    help="Usa OpenAI web_search para identificar partidos ya jugados y poblar sus ganadores.",
                ):
                    try:
                        with st.spinner("Buscando resultados de eliminatorias con OpenAI web_search..."):
                            updates = call_llm_knockout_results_update(model, fixtures_input)
                        merged = fixtures_input.copy()
                        if not updates.empty:
                            update_map = dict(zip(updates["match_id"], updates["winner"]))
                            merged["winner"] = merged["match_id"].map(update_map).fillna(merged["winner"])
                        normalized_knockout = normalize_knockout_results_update(merged, fixtures_input)
                        st.session_state["post_group_knockout_input"] = merged
                        st.session_state["post_group_knockout_results"] = normalized_knockout
                        clear_post_group_outputs()
                        save_post_group_state_from_session()
                        st.success("Resultados de eliminatorias actualizados con busqueda web y guardados en post_group_state.xlsx.")
                        st.rerun()
                    except Exception as exc:
                        st.error(f"No se pudieron actualizar las eliminatorias: {exc}")
            else:
                st.caption("OPENAI_API_KEY no detectada. Puedes cargar ganadores manualmente.")

            winner_options = [""]
            known_teams = set(fixtures_input["team_a"].astype(str).str.strip().tolist())
            known_teams.update(fixtures_input["team_b"].astype(str).str.strip().tolist())
            winner_options.extend(sorted(team for team in known_teams if team and not team.startswith("Ganador ")))

            edited_fixtures = st.data_editor(
                fixtures_input,
                width="stretch",
                num_rows="fixed",
                key="post_group_knockout_editor",
                disabled=["round", "match_id", "team_a", "team_b"],
                column_config={
                    "round": st.column_config.TextColumn("Ronda"),
                    "match_id": st.column_config.NumberColumn("Partido"),
                    "team_a": st.column_config.TextColumn("Equipo A"),
                    "team_b": st.column_config.TextColumn("Equipo B"),
                    "winner": st.column_config.SelectboxColumn(
                        "Ganador (si ya se jugo)",
                        options=winner_options,
                        required=False,
                    ),
                },
            )
            st.session_state["post_group_knockout_input"] = edited_fixtures.copy()

            if st.button(
                "Limpiar resultados KO cargados",
                help="Borra los ganadores registrados para volver a simular todas las llaves pendientes.",
            ):
                cleared = fixtures_input.copy()
                cleared["winner"] = ""
                st.session_state["post_group_knockout_input"] = cleared
                st.session_state["post_group_knockout_results"] = pd.DataFrame(columns=["match_id", "winner"])
                st.success("Se limpiaron los resultados de eliminatorias cargados.")
                st.rerun()

    if st.button(
        "Guardar resultados reales en post_group_state.xlsx",
        help=(
            "Persiste la tabla de grupos, fixtures de eliminatorias y ganadores "
            "actuales como fuente de verdad en resultados/post_group_state.xlsx."
        ),
        disabled=normalized_groups is None,
    ):
        try:
            fixtures_to_save = st.session_state.get("post_group_knockout_input", pd.DataFrame())
            if fixtures_to_save is None or fixtures_to_save.empty:
                knockout_results_to_save = pd.DataFrame(columns=["match_id", "winner"])
            else:
                knockout_results_to_save = normalize_knockout_results_update(
                    fixtures_to_save,
                    fixtures_to_save,
                )
            st.session_state["post_group_knockout_results"] = knockout_results_to_save
            save_post_group_state_from_session()
            st.success(f"Estado post-grupos guardado en {POST_GROUP_STATE_PATH}.")
        except Exception as exc:
            st.error(f"No se pudo guardar el estado post-grupos: {exc}")

    if st.button(
        "Modelar llaves con estos resultados",
        type="primary",
        help="Simula eliminatorias usando la tabla de grupos y, si existen, resultados reales de KO ya jugados.",
        disabled=normalized_groups is None,
    ):
        try:
            if normalized_groups is None:
                raise ValueError(
                    "La tabla de grupos tiene errores. Corrigela antes de modelar eliminatorias."
                )

            fixtures_for_validation = st.session_state.get("post_group_knockout_input", pd.DataFrame())
            if fixtures_for_validation is None or fixtures_for_validation.empty:
                knockout_results = pd.DataFrame(columns=["match_id", "winner"])
            else:
                knockout_results = normalize_knockout_results_update(
                    fixtures_for_validation,
                    fixtures_for_validation,
                )

            csv_text = dataframe_to_csv_text(df)
            group_csv_text = dataframe_to_csv_text(normalized_groups)
            knockout_csv_text = dataframe_to_csv_text(knockout_results)
            if saved_post_group_state is not None:
                fixture_override_csv_text = dataframe_to_csv_text(fixtures_for_validation)
            else:
                fixture_override_csv_text = dataframe_to_csv_text(
                    pd.DataFrame(columns=["match_id", "team_a", "team_b"])
                )
            with st.spinner("Simulando eliminatorias desde grupos cargados..."):
                st.session_state["post_group_projection"] = run_post_group_projection_cached(
                    csv_text,
                    group_csv_text,
                    knockout_csv_text,
                    fixture_override_csv_text,
                    params,
                )
                st.session_state["post_group_bracket"] = run_post_group_bracket_cached(
                    csv_text,
                    group_csv_text,
                    knockout_csv_text,
                    fixture_override_csv_text,
                    params,
                )
                st.session_state["post_group_bracket_probable"] = run_post_group_bracket_probable_cached(
                    csv_text,
                    group_csv_text,
                    knockout_csv_text,
                    fixture_override_csv_text,
                    params,
                )
                st.session_state["post_group_knockout_results"] = knockout_results
            st.success("Llaves modeladas desde los resultados de grupos cargados.")
        except Exception as exc:
            st.error(f"No se pudieron modelar las llaves: {exc}")

    projection = st.session_state.get("post_group_projection")
    bracket = st.session_state.get("post_group_bracket")
    bracket_probable = st.session_state.get("post_group_bracket_probable")
    if projection is None or not show_outputs:
        return

    favorite = projection.iloc[0]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Favorito post-grupos", favorite["team"], f"{favorite['champion_pct']:.1f}% campeon")
    c2.metric("Final", favorite["team"], f"{favorite['final_pct']:.1f}%")
    c3.metric("Semifinal", favorite["team"], f"{favorite['semifinal_pct']:.1f}%")
    c4.metric("Rating modelo", f"{favorite['overall']:.1f}", "0-100 relativo")

    col_mode, col_round = st.columns([2, 2])
    with col_mode:
        mode_options = []
        if bracket_probable is not None:
            mode_options.append("Mas probable post-grupos")
        if bracket is not None:
            mode_options.append("Una simulacion post-grupos")
        mode = st.radio("Modo post-grupos", mode_options, horizontal=True, key="post_group_bracket_mode")
    with col_round:
        start_round = st.radio(
            "Mostrar post-grupos desde",
            ["Ronda de 32", "Octavos de Final", "Cuartos de Final"],
            horizontal=True,
            key="post_group_bracket_start_round",
        )

    active = bracket_probable if mode == "Mas probable post-grupos" else bracket
    from_round = {
        "Ronda de 32": "round_of_32",
        "Octavos de Final": "round_of_16",
        "Cuartos de Final": "quarterfinal",
    }[start_round]
    if active is not None:
        fig = _build_bracket_figure(active, from_round)
        st.plotly_chart(fig, width="stretch")

    qualifiers = projection.loc[projection["round_of_32_pct"] > 0].copy()
    st.dataframe(
        qualifiers[
            [
                "team",
                "group",
                "overall",
                "round_of_32_pct",
                "round_of_16_pct",
                "quarterfinal_pct",
                "semifinal_pct",
                "final_pct",
                "champion_pct",
            ]
        ].round(2),
        width="stretch",
        hide_index=True,
    )


def render_app() -> None:
    """Orquestar el flujo principal de la aplicacion Streamlit.

    Llama a todos los componentes de renderizado en orden: barra lateral,
    titulo, editor de datos, ejecucion de la simulacion y las cuatro
    pestanas de contenido (Prediccion, Grupos, Modelo, LLM).

    Returns
    -------
    None
        Renderiza la aplicacion completa y actualiza ``st.session_state``.
    """
    load_persisted_outputs_once()

    params, llm_model = render_sidebar()
    st.session_state["params"] = params

    st.title(APP_TITLE)
    st.caption(APP_CAPTION)

    default_df = load_default_data_cached()
    working_df = render_data_editor(default_df, llm_model)

    st.subheader("Alcance de simulacion")
    simulation_scope = st.radio(
        "Que quieres modelar?",
        ["Torneo completo", "Desde resultados de grupos"],
        horizontal=True,
        key="simulation_scope",
        help="Elige si quieres simular fase de grupos + eliminatorias, o fijar la tabla final de grupos y simular solo las llaves.",
    )
    post_group_mode = simulation_scope == "Desde resultados de grupos"

    if post_group_mode:
        st.caption("Flujo recomendado: 1) validar grupos, 2) cargar KO jugados, 3) modelar restantes.")
    else:
        st.caption("Flujo recomendado: ajustar ratings/parametros y ejecutar simulacion completa.")

    if not post_group_mode:
        if st.button("Simular torneo completo", type="primary", help="Ejecuta fase de grupos y eliminatorias con los ratings y parametros actuales."):
            try:
                csv_text = dataframe_to_csv_text(working_df)
                run_simulation_cached.clear()
                run_bracket_sample_cached.clear()
                run_bracket_probable_cached.clear()
                with st.spinner("Simulando torneos..."):
                    st.session_state["simulation_results"] = run_simulation_cached(csv_text, params)
                with st.spinner("Generando cuadro de eliminacion..."):
                    st.session_state["bracket"] = run_bracket_sample_cached(csv_text, params)
                with st.spinner("Calculando cuadro mas probable (1000 simulaciones)..."):
                    st.session_state["bracket_probable"] = run_bracket_probable_cached(csv_text, params)
                st.session_state["simulation_df"] = working_df.copy()
                save_montecarlo_results(
                    st.session_state["simulation_results"],
                    st.session_state["simulation_df"],
                    params,
                    bracket=st.session_state["bracket"],
                    bracket_probable=st.session_state["bracket_probable"],
                )
            except Exception as exc:
                st.error(f"No se pudo correr la simulacion: {exc}")
    else:
        render_post_group_knockout_view(
            working_df,
            params,
            llm_model,
            show_header=False,
            show_outputs=False,
        )

    if post_group_mode:
        results = st.session_state.get("post_group_projection")
        sim_df = working_df
        bracket = st.session_state.get("post_group_bracket")
        bracket_probable = st.session_state.get("post_group_bracket_probable")
    else:
        results = st.session_state.get("simulation_results")
        sim_df = st.session_state.get("simulation_df", working_df)
        bracket = st.session_state.get("bracket")
        bracket_probable = st.session_state.get("bracket_probable")

    tab_pred, tab_bracket, tab_groups, tab_model, tab_llm, tab_eval, tab_report = st.tabs(
        ["Prediccion", "Bracket", "Grupos", "Modelo", "LLM", "Evaluar", "Reporte"]
    )
    with tab_pred:
        if results is None:
            if post_group_mode:
                st.info("Carga los resultados de grupos y pulsa **Modelar llaves con estos grupos** para ver las predicciones.")
            else:
                st.info("Pulsa **Simular torneo completo** para ver las predicciones.")
        else:
            render_probability_view(results)
    with tab_bracket:
        if post_group_mode:
            render_bracket_view(
                bracket,
                bracket_probable,
                empty_message="Carga los resultados de grupos y pulsa **Modelar llaves con estos grupos** para generar el cuadro de eliminacion.",
                probable_label="Mas probable post-grupos",
                sample_label="Una simulacion post-grupos",
                probable_caption="Cuadro mas probable condicionado a los resultados de grupos cargados.",
                sample_caption="Simulacion representativa condicionada a los resultados de grupos cargados.",
            )
        else:
            render_bracket_view(bracket, bracket_probable)
    with tab_groups:
        if post_group_mode:
            st.info("En este modo la fase de grupos ya esta fijada por la tabla que cargaste arriba.")
            st.dataframe(
                st.session_state.get("post_group_results_input", pd.DataFrame()),
                width="stretch",
                hide_index=True,
            )
        elif results is None:
            st.info("Pulsa **Simular torneo completo** para ver el analisis por grupos.")
        else:
            render_group_view(sim_df, results)
    with tab_model:
        render_model_view()
    with tab_llm:
        if results is None:
            render_llm_view(None, sim_df, llm_model, bracket_probable)
        else:
            render_llm_view(results, sim_df, llm_model, bracket_probable)
    with tab_eval:
        render_evaluation_view()
    with tab_report:
        render_report_view(results, sim_df, bracket_probable)


def render_evaluation_view() -> None:
    """Renderizar la pestana de evaluacion de submissions.
    
    Muestra controles para validar los CSVs de ``evaluations/`` bajo
    demanda, calcular Brier Scores cuando existe ``ground_truth.csv`` y
    generar un reporte Excel descargable.

    Returns
    -------
    None
        Renderiza controles, tablas, graficos y botones de descarga en
        la pestana de Streamlit.
    """
    st.markdown("### Evaluación de Submissions")
    st.markdown("Valida y evalúa todos los CSVs de la carpeta `evaluations/`")
    st.markdown("")
    
    # Definir equipos válidos y columnas
    valid_teams = [
        "Algeria", "Argentina", "Australia", "Austria", "Belgium",
        "Bosnia and Herzegovina", "Brazil", "Cabo Verde", "Canada", "Colombia",
        "Congo DR", "Cote d'Ivoire", "Croatia", "Curacao", "Czechia",
        "Ecuador", "Egypt", "England", "France", "Germany",
        "Ghana", "Haiti", "IR Iran", "Iraq", "Japan",
        "Jordan", "Korea Republic", "Mexico", "Morocco", "Netherlands",
        "New Zealand", "Norway", "Panama", "Paraguay", "Portugal",
        "Qatar", "Saudi Arabia", "Scotland", "Senegal", "South Africa",
        "Spain", "Sweden", "Switzerland", "Tunisia", "Turkiye",
        "USA", "Uzbekistan", "Uruguay",
    ]
    
    required_columns = ["team", "prob_champion", "prob_final", "prob_semifinal"]
    stage_weights = {
        "prob_champion": 0.50,
        "prob_final": 0.30,
        "prob_semifinal": 0.20,
    }
    
    def calculate_brier_score(predictions: pd.Series, actuals: pd.Series) -> float:
        """Calcular el Brier Score de una serie de predicciones.

        Parameters
        ----------
        predictions : pd.Series
            Probabilidades predichas para una etapa del torneo.
        actuals : pd.Series
            Indicadores reales de ocurrencia, codificados como 0 o 1.

        Returns
        -------
        float
            Promedio de ``(prediccion - real) ** 2``.
        """
        return ((predictions - actuals) ** 2).mean()
    
    def validate_submission(df: pd.DataFrame, filename: str) -> tuple[bool, list[str]]:
        """Validar que una submission tenga el formato esperado.

        Parameters
        ----------
        df : pd.DataFrame
            Datos cargados desde el CSV de submission.
        filename : str
            Nombre del archivo, usado para contexto en mensajes.

        Returns
        -------
        tuple[bool, list[str]]
            Bandera de validez y lista de errores encontrados.
        """
        errors = []
        
        # Verificar columnas
        missing_cols = set(required_columns) - set(df.columns)
        if missing_cols:
            errors.append(f"Faltan columnas: {', '.join(missing_cols)}")
        
        extra_cols = set(df.columns) - set(required_columns)
        if extra_cols:
            errors.append(f"Columnas extra: {', '.join(extra_cols)}")
        
        if errors:
            return False, errors
        
        # Verificar cantidad de equipos
        if len(df) != 48:
            errors.append(f"Se esperan 48 equipos, se encontraron {len(df)}")
        
        # Verificar nombres de equipos
        invalid_teams = set(df["team"]) - set(valid_teams)
        if invalid_teams:
            errors.append(f"Equipos inválidos: {', '.join(sorted(invalid_teams))}")
        
        missing_teams = set(valid_teams) - set(df["team"])
        if missing_teams:
            errors.append(f"Equipos faltantes: {', '.join(sorted(missing_teams))}")
        
        # Verificar duplicados
        if df["team"].duplicated().any():
            dups = df.loc[df["team"].duplicated(), "team"].tolist()
            errors.append(f"Equipos duplicados: {', '.join(dups)}")
        
        # Verificar rangos de probabilidades
        for col in required_columns[1:]:
            invalid = ((df[col] < 0) | (df[col] > 1)).sum()
            if invalid > 0:
                errors.append(f"Columna {col}: {invalid} valores fuera de rango [0, 1]")
        
        return len(errors) == 0, errors
    
    def evaluate_submission(csv_path: Path, ground_truth: pd.DataFrame) -> dict:
        """Validar y puntuar un archivo de submission.

        Parameters
        ----------
        csv_path : Path
            Ruta del archivo CSV a evaluar.
        ground_truth : pd.DataFrame
            Resultados reales del torneo con columnas ``team``,
            ``champion``, ``final`` y ``semifinal``.

        Returns
        -------
        dict
            Resultado de validacion y scores Brier por etapa, incluyendo
            ``score_final`` cuando hay datos reales disponibles.
        """
        filename = csv_path.name
        result = {
            "file": filename,
            "valid": False,
            "errors": [],
            "brier_champion": None,
            "brier_final": None,
            "brier_semifinal": None,
            "score_final": None,
        }
        
        try:
            df = pd.read_csv(csv_path)
        except Exception as e:
            result["errors"].append(f"Error al leer CSV: {str(e)}")
            return result
        
        valid, errors = validate_submission(df, filename)
        result["valid"] = valid
        result["errors"] = errors
        
        if not valid or ground_truth.empty:
            return result
        
        # Merge con resultados reales
        merged = df.merge(ground_truth, on="team", how="inner")
        
        if len(merged) != 48:
            result["errors"].append("No se pudieron hacer match todos los equipos")
            return result
        
        # Calcular Brier Scores por etapa
        result["brier_champion"] = calculate_brier_score(
            merged["prob_champion"], merged["champion"]
        )
        result["brier_final"] = calculate_brier_score(
            merged["prob_final"], merged["final"]
        )
        result["brier_semifinal"] = calculate_brier_score(
            merged["prob_semifinal"], merged["semifinal"]
        )
        
        # Calcular score final ponderado
        result["score_final"] = (
            stage_weights["prob_champion"] * result["brier_champion"]
            + stage_weights["prob_final"] * result["brier_final"]
            + stage_weights["prob_semifinal"] * result["brier_semifinal"]
        )
        
        return result

    def build_excel_report(results: list[dict], ground_truth_loaded: bool, eval_dir: Path) -> bytes:
        """Crear un reporte XLSX con formato para descargar desde Streamlit.

        Parameters
        ----------
        results : list[dict]
            Resultados producidos por ``evaluate_submission``.
        ground_truth_loaded : bool
            Indica si ``ground_truth.csv`` fue cargado correctamente.
        eval_dir : Path
            Carpeta donde se buscaron las submissions.

        Returns
        -------
        bytes
            Contenido binario del archivo XLSX generado en memoria.
        """
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        validation_df = pd.DataFrame(
            [
                {
                    "Submission": result["file"],
                    "Estado": "Valido" if result["valid"] else "Error",
                    "Detalles": "; ".join(result["errors"]) if result["errors"] else "OK",
                }
                for result in results
            ]
        )
        ranked_results = sorted(
            [r for r in results if r["valid"] and r["score_final"] is not None],
            key=lambda x: x["score_final"],
        )
        scores_df = pd.DataFrame(
            [
                {
                    "Posicion": i,
                    "Submission": result["file"],
                    "Champion": result["brier_champion"],
                    "Final": result["brier_final"],
                    "Semifinal": result["brier_semifinal"],
                    "Score Final": result["score_final"],
                }
                for i, result in enumerate(ranked_results, 1)
            ]
        )
        best = ranked_results[0] if ranked_results else None
        valid_count = sum(1 for result in results if result["valid"])
        summary_df = pd.DataFrame(
            [
                {"Metrica": "Carpeta evaluada", "Valor": str(eval_dir)},
                {"Metrica": "Submissions encontradas", "Valor": len(results)},
                {"Metrica": "Submissions validas", "Valor": valid_count},
                {"Metrica": "Submissions con error", "Valor": len(results) - valid_count},
                {"Metrica": "Ground truth cargado", "Valor": "Si" if ground_truth_loaded else "No"},
                {"Metrica": "Mejor submission", "Valor": best["file"] if best else "N/A"},
                {"Metrica": "Mejor score final", "Valor": best["score_final"] if best else "N/A"},
            ]
        )

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Resumen", index=False)
            validation_df.to_excel(writer, sheet_name="Validacion", index=False)
            if not scores_df.empty:
                scores_df.to_excel(writer, sheet_name="Ranking", index=False)

            workbook = writer.book
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            summary_fill = PatternFill("solid", fgColor="D9EAF7")
            ok_fill = PatternFill("solid", fgColor="E2F0D9")
            error_fill = PatternFill("solid", fgColor="FCE4D6")
            podium_fills = {
                1: PatternFill("solid", fgColor="FFF2CC"),
                2: PatternFill("solid", fgColor="D9EAF7"),
                3: PatternFill("solid", fgColor="EADCF8"),
            }

            for worksheet in workbook.worksheets:
                worksheet.freeze_panes = "A2"
                worksheet.sheet_view.showGridLines = False
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                for column_cells in worksheet.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column_cells)
                    width = min(max(max_length + 3, 14), 55)
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

            resumen = workbook["Resumen"]
            for row in resumen.iter_rows(min_row=2, max_col=2):
                row[0].font = Font(bold=True)
                row[0].fill = summary_fill

            validacion = workbook["Validacion"]
            for row in validacion.iter_rows(min_row=2):
                fill = ok_fill if row[1].value == "Valido" else error_fill
                for cell in row:
                    cell.fill = fill

            if "Ranking" in workbook.sheetnames:
                ranking = workbook["Ranking"]
                for row in ranking.iter_rows(min_row=2):
                    fill = podium_fills.get(row[0].value)
                    if fill:
                        for cell in row:
                            cell.fill = fill
                    for cell in row[2:]:
                        cell.number_format = "0.000000"

                chart = BarChart()
                chart.title = "Score final por submission"
                chart.y_axis.title = "Score final"
                chart.x_axis.title = "Submission"
                data = Reference(ranking, min_col=6, min_row=1, max_row=ranking.max_row)
                categories = Reference(ranking, min_col=2, min_row=2, max_row=ranking.max_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.height = 8
                chart.width = 18
                ranking.add_chart(chart, "H2")

        return output.getvalue()
    
    # Encontrar archivos
    eval_dir = Path(__file__).resolve().parents[1] / "evaluations"
    
    if not eval_dir.exists():
        st.error(f"❌ Carpeta evaluations/ no encontrada en {eval_dir}")
        return
    
    csv_files = sorted([f for f in eval_dir.glob("*.csv") if f.name != "ground_truth.csv"])

    if not csv_files:
        st.warning("No se encontraron archivos CSV para evaluar.")
        return

    ground_truth_path = eval_dir / "ground_truth.csv"
    has_ground_truth = ground_truth_path.exists()
    st.info(
        f"{len(csv_files)} submission(s) encontrada(s). "
        f"Ground truth: {'disponible' if has_ground_truth else 'no encontrado'}."
    )

    run_col, report_col, download_col = st.columns([1, 1, 1])

    with run_col:
        evaluate_clicked = st.button(
            "Evaluar submissions",
            type="primary",
            help="Valida los CSVs y calcula Brier Scores si existe ground_truth.csv.",
        )

    if evaluate_clicked:
        ground_truth = pd.DataFrame()
        ground_truth_loaded = False
        if has_ground_truth:
            try:
                ground_truth = pd.read_csv(ground_truth_path)
                ground_truth_loaded = True
            except Exception as e:
                st.warning(f"Error al cargar ground_truth.csv: {e}")
        else:
            st.warning("Archivo ground_truth.csv no encontrado.")

        with st.spinner("Evaluando submissions..."):
            st.session_state["evaluation_results"] = [
                evaluate_submission(csv_path, ground_truth) for csv_path in csv_files
            ]
            st.session_state["evaluation_ground_truth_loaded"] = ground_truth_loaded
            st.session_state.pop("evaluation_excel_report", None)

    results = st.session_state.get("evaluation_results")
    ground_truth_loaded = st.session_state.get("evaluation_ground_truth_loaded", False)

    with report_col:
        generate_report_clicked = st.button(
            "Generar reporte Excel",
            disabled=not results,
            help="Crea un XLSX con resumen, validacion, ranking y grafico.",
        )

    if generate_report_clicked and results:
        with st.spinner("Generando reporte Excel..."):
            st.session_state["evaluation_excel_report"] = build_excel_report(
                results,
                ground_truth_loaded,
                eval_dir,
            )
        st.success("Reporte Excel listo.")

    with download_col:
        report_bytes = st.session_state.get("evaluation_excel_report")
        st.download_button(
            "Descargar reporte",
            data=report_bytes or b"",
            file_name="reporte_evaluacion_wcup2026.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=report_bytes is None,
            help="Descarga el reporte generado.",
        )

    if not results:
        st.info("Pulsa **Evaluar submissions** para ver la validacion y los scores.")
        return

    st.divider()

    st.markdown("#### Validacion de Formato")
    validation_df = pd.DataFrame(
        [
            {
                "Submission": result["file"],
                "Estado": "Valido" if result["valid"] else "Error",
                "Detalles": "; ".join(result["errors"]) if result["errors"] else "OK",
            }
            for result in results
        ]
    )
    st.dataframe(validation_df, width="stretch", hide_index=True)

    st.divider()

    if ground_truth_loaded:
        st.markdown("#### Resultados de Evaluacion (Brier Score)")
        st.markdown("*Ordenado por Score Final (menor es mejor)*")

        results_sorted = sorted(
            [r for r in results if r["valid"] and r["score_final"] is not None],
            key=lambda x: x["score_final"],
        )

        if results_sorted:
            scores_df = pd.DataFrame(
                [
                    {
                        "Posicion": i,
                        "Submission": result["file"],
                        "Champion": result["brier_champion"],
                        "Final": result["brier_final"],
                        "Semifinal": result["brier_semifinal"],
                        "Score Final": result["score_final"],
                    }
                    for i, result in enumerate(results_sorted, 1)
                ]
            )

            def highlight_row(row):
                """Asignar color de fondo a las primeras posiciones.

                Parameters
                ----------
                row : pd.Series
                    Fila del DataFrame de ranking renderizado por
                    ``pandas.Styler``.

                Returns
                -------
                list[str]
                    Estilos CSS por celda para resaltar el podio.
                """
                if row["Posicion"] == 1:
                    return ["background-color: #90EE90"] * len(row)
                if row["Posicion"] == 2:
                    return ["background-color: #87CEEB"] * len(row)
                if row["Posicion"] == 3:
                    return ["background-color: #FFB6C1"] * len(row)
                return [""] * len(row)

            st.dataframe(
                scores_df.style.format({
                    "Champion": "{:.6f}",
                    "Final": "{:.6f}",
                    "Semifinal": "{:.6f}",
                    "Score Final": "{:.6f}",
                }).apply(highlight_row, axis=1),
                width="stretch",
                hide_index=True,
            )

            st.markdown("#### Comparacion de Scores Finales")
            plot_scores = scores_df.sort_values("Score Final", ascending=False).tail(15)
            breakdown = scores_df.sort_values("Score Final").head(10)
            fig = make_subplots(
                rows=1,
                cols=2,
                subplot_titles=("Ranking final (menor es mejor)", "Desglose por etapa del top 10"),
                column_widths=[0.45, 0.55],
            )
            fig.add_trace(
                go.Bar(
                    x=plot_scores["Score Final"],
                    y=plot_scores["Submission"],
                    orientation="h",
                    marker={
                        "color": plot_scores["Score Final"],
                        "colorscale": [[0, "#166534"], [0.55, "#facc15"], [1, "#b91c1c"]],
                        "line": {"color": "rgba(31,41,55,0.25)", "width": 0.7},
                    },
                    text=plot_scores["Score Final"].map(lambda value: f"{value:.5f}"),
                    textposition="outside",
                    hovertemplate="<b>%{y}</b><br>Score final: %{x:.6f}<extra></extra>",
                    showlegend=False,
                ),
                row=1,
                col=1,
            )
            stage_colors = {
                "Champion": "#14532d",
                "Final": "#2563eb",
                "Semifinal": "#f59e0b",
            }
            for stage in ["Champion", "Final", "Semifinal"]:
                fig.add_trace(
                    go.Bar(
                        x=breakdown["Submission"],
                        y=breakdown[stage],
                        name=stage,
                        marker={"color": stage_colors[stage]},
                        hovertemplate=f"<b>%{{x}}</b><br>{stage}: %{{y:.6f}}<extra></extra>",
                    ),
                    row=1,
                    col=2,
                )
            fig.update_layout(
                title="Calibracion de submissions",
                barmode="group",
                legend={"orientation": "h", "y": -0.22},
            )
            fig.update_xaxes(title="Score final", row=1, col=1)
            fig.update_yaxes(title="", row=1, col=1)
            fig.update_xaxes(title="", tickangle=-35, row=1, col=2)
            fig.update_yaxes(title="Brier Score", row=1, col=2)
            fig.update_traces(cliponaxis=False)
            st.plotly_chart(apply_plotly_theme(fig, height=520), width="stretch")
        else:
            st.info("No hay submissions validas para mostrar scores.")
    else:
        st.info("Para ver scores de evaluacion, crea archivo `ground_truth.csv` en la carpeta `evaluations/`")
        st.markdown("**Estructura esperada:**")
        st.code("team,champion,final,semifinal\nArgentina,1,1,1\nFrance,0,1,1\n...", language="csv")

    return


def main() -> None:
    """Punto de entrada de la aplicacion Streamlit.

    Configura la pagina y lanza el renderizado completo de la app.  Es
    invocado por ``app.py`` al ejecutar ``streamlit run app.py``.

    Returns
    -------
    None
        Inicializa y renderiza la aplicacion.
    """
    configure_page()
    render_app()
